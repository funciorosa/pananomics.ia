"""
Carga los archivos Excel de PbR a Supabase.

Uso desde la raíz del proyecto:
    python scripts/load_pbr_data.py

Requiere:
    pip install openpyxl requests python-dotenv

Variables en .env.local (raíz del proyecto):
    SUPABASE_URL=https://eyxgyeybvokvrkrarmzh.supabase.co
    SUPABASE_SERVICE_ROLE_KEY=<tu service role key>

Los Excel deben estar en: C:\\Users\\mrmma\\Desktop\\Pananomics\\PbR
Nombre de archivo: 005_MIRE_PbR_Supabase.xlsx, 010_MIDA_PbR_Supabase.xlsx, etc.
Fila 3 = headers, fila 4 en adelante = datos.
"""

import os
import sys
import json
from pathlib import Path

import openpyxl
import requests
from dotenv import load_dotenv

# ── Configuración ─────────────────────────────────────────────────────────────
load_dotenv(Path(__file__).parent.parent / ".env.local")

SUPABASE_URL = os.environ.get("SUPABASE_URL", "").rstrip("/")
SUPABASE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")

if not SUPABASE_URL or not SUPABASE_KEY:
    print("❌  Faltan variables en .env.local:")
    print("    SUPABASE_URL=https://eyxgyeybvokvrkrarmzh.supabase.co")
    print("    SUPABASE_SERVICE_ROLE_KEY=<tu service role key>")
    sys.exit(1)

CARPETA = Path(r"C:\Users\mrmma\Desktop\Pananomics\PbR")

HEADERS = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "resolution=merge-duplicates,return=minimal",
}

CONFLICT_KEYS = {
    "pbr_entidades":          "codigo_entidad",
    "pbr_programas":          "codigo_entidad,anio,codigo_programa",
    "pbr_plurianual":         "codigo_entidad,codigo_programa,anio,fuente_anexo",
    "pbr_ejes_estrategicos":  "codigo_entidad,numero_eje",
}

HOJAS_TABLAS = [
    "pbr_entidades",
    "pbr_programas",
    "pbr_plurianual",
    "pbr_subprogramas",
    "pbr_ods",
    "pbr_ejes_estrategicos",
    "pbr_equivalencia",
]

# ── Helpers ───────────────────────────────────────────────────────────────────

def leer_hoja(ws):
    """Lee una hoja Excel: fila 3 = headers, fila 4+ = datos."""
    headers = [c.value for c in ws[3]]
    rows = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not any(v is not None for v in row):
            continue
        rec = {}
        for h, v in zip(headers, row):
            if not h:
                continue
            if isinstance(v, str):
                v = v.strip() or None
            if v == "TRUE":
                v = True
            elif v == "FALSE":
                v = False
            rec[str(h).strip()] = v
        # Quitar campos auto-generados
        rec.pop("id", None)
        rec.pop("created_at", None)
        # Solo agregar si hay al menos un valor no nulo
        if any(val is not None for val in rec.values()):
            rows.append(rec)
    return rows


def procesar_arrays(rec):
    """Convierte '{A,B,C}' → ['A','B','C'] en columnas de tipo array."""
    for col in ("pilares_peg_2025", "pilares_peg_2019"):
        val = rec.get(col)
        if isinstance(val, str):
            limpio = val.strip().strip("{}")
            rec[col] = [x.strip() for x in limpio.split(",") if x.strip()] if limpio else []
    return rec


def upsert(tabla, rows):
    """Upsert batch a Supabase REST API."""
    if not rows:
        return 0

    conflict = CONFLICT_KEYS.get(tabla)
    url = f"{SUPABASE_URL}/rest/v1/{tabla}"
    params = {"on_conflict": conflict} if conflict else {}

    # Enviar en lotes de 200 para evitar payload demasiado grande
    BATCH = 200
    total = 0
    for i in range(0, len(rows), BATCH):
        batch = rows[i : i + BATCH]
        res = requests.post(url, headers=HEADERS, params=params, data=json.dumps(batch, default=str))
        if res.status_code not in (200, 201, 204):
            raise RuntimeError(f"HTTP {res.status_code}: {res.text[:300]}")
        total += len(batch)
    return total


def insertar(tabla, rows):
    """Insert simple (para tablas sin conflict key definida)."""
    if not rows:
        return 0
    url = f"{SUPABASE_URL}/rest/v1/{tabla}"
    res = requests.post(url, headers=HEADERS, data=json.dumps(rows, default=str))
    if res.status_code not in (200, 201, 204):
        raise RuntimeError(f"HTTP {res.status_code}: {res.text[:300]}")
    return len(rows)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    if not CARPETA.exists():
        print(f"❌  Carpeta no encontrada: {CARPETA}")
        sys.exit(1)

    archivos = sorted(CARPETA.glob("*.xlsx"))
    if not archivos:
        print(f"⚠️  No se encontraron archivos .xlsx en {CARPETA}")
        sys.exit(0)

    print(f"📂  {len(archivos)} archivo(s) encontrado(s) en {CARPETA}\n")

    total_ok = 0
    total_err = 0

    for archivo in archivos:
        print(f"── {archivo.name}")
        try:
            wb = openpyxl.load_workbook(archivo, data_only=True)
        except Exception as e:
            print(f"   ❌ No se pudo abrir: {e}")
            total_err += 1
            continue

        for tabla in HOJAS_TABLAS:
            if tabla not in wb.sheetnames:
                continue
            try:
                rows = [procesar_arrays(r) for r in leer_hoja(wb[tabla])]
                if not rows:
                    print(f"   {tabla}: (vacío, se omite)")
                    continue

                if tabla in CONFLICT_KEYS:
                    n = upsert(tabla, rows)
                else:
                    # Para tablas sin conflict key, hacemos delete+insert por entidad
                    # Detectamos el codigo_entidad del primer registro
                    codigo = rows[0].get("codigo_entidad")
                    if codigo:
                        enc = requests.utils.quote(codigo, safe="")
                        requests.delete(
                            f"{SUPABASE_URL}/rest/v1/{tabla}",
                            headers=HEADERS,
                            params={"codigo_entidad": f"eq.{codigo}"},
                        )
                    n = insertar(tabla, rows)

                print(f"   ✅ {tabla}: {n} filas")
                total_ok += 1
            except Exception as e:
                print(f"   ❌ {tabla}: {e}")
                total_err += 1

        wb.close()
        print()

    print("─" * 50)
    print(f"✅ Tablas procesadas OK: {total_ok}")
    if total_err:
        print(f"❌ Errores: {total_err}")
    print("Carga completa.")


if __name__ == "__main__":
    main()
